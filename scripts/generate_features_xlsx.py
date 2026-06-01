#!/usr/bin/env python3
"""Generate XLSX workbook: Feature Tracking + Test Cases for AutoBBoxToolkit."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_REPO = Path(".")
DEFAULT_OUTPUT = Path("docs/AutoBBoxToolkit_Features.xlsx")

HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
PENDING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def apply_header(ws, headers: list[str], row: int = 1):
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def apply_cell(ws, row: int, col: int, value: Any, fill=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = CELL_ALIGN
    cell.border = THIN_BORDER
    if fill:
        cell.fill = fill
    return cell


def build_feature_sheet(wb: Workbook, repo: Path):
    ws = wb.active
    ws.title = "功能清单 Feature List"

    headers = ["功能ID", "功能名称", "命令", "源文件数", "状态", "备注"]
    apply_header(ws, headers)

    fi = load_json(repo / ".autobbox/index/feature_index.json", {"features": []})
    features = fi.get("features", [])

    for i, feat in enumerate(features):
        row = i + 2
        fid = feat.get("feature_id", "")
        title = feat.get("title", "")
        cmds = ", ".join(feat.get("commands", [])[:4])
        paths = feat.get("paths", {})
        path_count = sum(len(v) if isinstance(v, list) else 1 for v in paths.values())
        notes = feat.get("notes", "")

        apply_cell(ws, row, 1, fid)
        apply_cell(ws, row, 2, title)
        apply_cell(ws, row, 3, cmds)
        apply_cell(ws, row, 4, path_count)
        apply_cell(ws, row, 5, "✅ 已实现", PENDING_FILL)
        apply_cell(ws, row, 6, notes)

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 30

    ws.auto_filter.ref = f"A1:F{len(features) + 1}"
    ws.freeze_panes = "A2"


def build_test_sheet(wb: Workbook):
    ws = wb.create_sheet("测试用例 Test Cases")

    headers = ["测试ID", "功能模块", "测试场景", "前置条件", "操作步骤", "预期结果", "状态", "备注"]
    apply_header(ws, headers)

    test_cases = [
        # BOM
        ("TC-BOM-01", "BOM清单", "打开装配后查看BOM", "装配中打开模型", "点击BOM清单按钮", "显示BOM表格,列出所有组件及参数", "待验证", ""),
        ("TC-BOM-02", "BOM清单", "BOM模型名称过滤", "装配中打开BOM对话框", "在模型过滤器输入关键词按Enter", "BOM表格仅显示匹配行", "待验证", ""),
        ("TC-BOM-03", "BOM清单", "BOM参数名称过滤", "装配中打开BOM对话框", "在参数过滤器输入关键词按Enter", "BOM表格仅显示匹配行", "待验证", ""),
        ("TC-BOM-04", "BOM清单", "BOM组件/零件勾选过滤", "装配中打开BOM对话框", "取消勾选'零件'复选框", "仅显示组件行", "待验证", ""),
        ("TC-BOM-05", "BOM清单", "添加自定义参数列", "BOM对话框可见参数列表", "输入参数名+类型+默认值,点添加", "新参数列出现在BOM表格中", "待验证", ""),
        ("TC-BOM-06", "BOM清单", "更新参数到模型", "BOM中有勾选的草稿值", "点'更新到模型'", "参数值写入对应模型", "待验证", ""),
        ("TC-BOM-07", "BOM清单", "导出Excel", "BOM对话框打开", "点'导出Excel'", "在工作目录生成CSV文件", "待验证", ""),

        # Quick Simprep
        ("TC-QS-01", "快速简化表示", "按PTC_COMMON_NAME分类创建", "装配中有带PTC_COMMON_NAME的直接子组件", "打开快速简化表示,勾选分类,点每类创建", "每个分类创建独立简化表示", "待验证", ""),
        ("TC-QS-02", "快速简化表示", "合并创建简化表示", "同上", "勾选多个分类,点合并创建", "创建单个合并简化表示", "待验证", ""),
        ("TC-QS-03", "快速简化表示", "模式切换", "打开快速简化表示对话框", "切换ModeMenu到'管理已有表示'", "显示已有简化表示列表", "待验证", ""),
        ("TC-QS-04", "快速简化表示", "添加到已有表示", "有已创建的简化表示+新分类", "在创建模式勾选分类,点'添加到已有表示'", "该分类组件加入目标简化表示", "待验证", ""),
        ("TC-QS-05", "快速简化表示", "删除表示中分类", "有包含多分类的简化表示", "管理模式选中rep,点'删除选中分类'", "该分类组件从rep中移除", "待验证", ""),
        ("TC-QS-06", "快速简化表示", "刷新分类", "装配中某分类的组件有变更", "管理模式选中rep,点'刷新选中分类'", "rep中该分类组件更新为最新", "待验证", ""),
        ("TC-QS-07", "快速简化表示", "重命名简化表示", "有已有简化表示", "管理模式选中rep,点'重命名'", "rep名称变更", "待验证", ""),

        # Smart Dimension
        ("TC-SD-01", "智能尺寸", "单参考创建尺寸", "工程图中选择视图边", "点击参考→放置尺寸", "创建尺寸标注", "待验证", ""),
        ("TC-SD-02", "智能尺寸", "切线边选择", "视图切线边显示样式为实线", "选择圆角/切线边→放置尺寸", "成功选择切线边并创建尺寸", "待验证", ""),

        # Drawing Views
        ("TC-DV-01", "建视图", "批量创建工程图视图", "工程图中", "勾选选项,点建视图", "批量创建视图", "待验证", ""),
        ("TC-DV-02", "视图整理", "自动排列视图", "有重叠视图的工程图", "框选视图,点视图整理", "视图自动拆组消除重叠", "待验证", ""),

        # Family Table
        ("TC-FT-01", "族表管理器", "打开族表管理", "装配中有族表实例", "点击族表管理器", "打开多层族表管理器对话框", "待验证", ""),
        ("TC-FT-02", "族表管理器", "Excel导出导入", "族表管理器打开", "点导出→修改→导入", "族表数据正确往返", "待验证", ""),

        # Random Color
        ("TC-RC-01", "随机上色", "装配随机上色", "装配打开", "点随机上色", "装配中模型随机着色", "待验证", ""),

        # Batch Rename
        ("TC-BR-01", "批量重命名", "批量修改模型名", "装配打开", "打开批量重命名,修改名称", "模型名称批量变更", "待验证", ""),
    ]

    for i, tc in enumerate(test_cases):
        row = i + 2
        for col, val in enumerate(tc, 1):
            fill = None
            if col == 7:  # Status column
                if val == "通过":
                    fill = PASS_FILL
                elif val == "失败":
                    fill = FAIL_FILL
                else:
                    fill = PENDING_FILL
            apply_cell(ws, row, col, val, fill)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 32
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 20

    ws.auto_filter.ref = f"A1:H{len(test_cases) + 1}"
    ws.freeze_panes = "A2"


def build_api_sheet(wb: Workbook, repo: Path):
    ws = wb.create_sheet("API参考 API Reference")

    headers = ["API名称", "类别", "概述", "返回值", "项目使用文件数"]
    apply_header(ws, headers)

    api_index = load_json(repo / "docs/api_extracts/api_quickref.json", {"apis": {}})
    apis = api_index.get("apis", {})

    # Load API graph for usage counts
    graph = load_json(repo / ".autobbox/index/api_graph.json", {})
    nodes = graph.get("nodes", {})

    row = 2
    for api_name in sorted(apis):
        info = apis[api_name]
        usage = nodes.get(api_name, {})
        file_count = usage.get("file_count", 0) if isinstance(usage, dict) else 0

        apply_cell(ws, row, 1, api_name)
        apply_cell(ws, row, 2, info.get("category", ""))
        apply_cell(ws, row, 3, info.get("synopsis", "")[:200])
        apply_cell(ws, row, 4, info.get("returns", "")[:100])
        apply_cell(ws, row, 5, file_count)
        row += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 14

    ws.auto_filter.ref = f"A1:E{row - 1}"
    ws.freeze_panes = "A2"


def main() -> int:
    repo = DEFAULT_REPO.resolve()
    output = DEFAULT_OUTPUT

    wb = Workbook()
    build_feature_sheet(wb, repo)
    build_test_sheet(wb)
    build_api_sheet(wb, repo)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    print(f"XLSX written: {output}")
    print(f"  Sheets: {wb.sheetnames}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
